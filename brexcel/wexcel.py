# coding=utf8
from openpyxl import Workbook


class WExcel:

    header_alias = {}
    header_order = []

    def __init__(self, data):
        self.data = data


    def SaveExcelAs(self, filename, sheet='Sheet1'):

        """
        :param filename:
        :param sheet:
        :return:
        """

        header_alias = {}

        for row in self.data:
            for colname in row:
                if colname not in self.header_alias:
                    header_alias[colname] = colname

        for colname in self.header_alias:
            header_alias[colname] = self.header_alias[colname]

        # if header_alias is empty, I fill it

        if self.header_alias != {}:
            for row in self.data:
                for colname in row:
                    if colname not in self.header_alias:
                        self.header_alias[colname] = colname

        # if header_order is empty, I fill it
        if self.header_order == []:
            for colname in self.header_alias:
                self.header_order.append(colname)

        if filename == None:
            raise Exception("filename not provided. What's the filename of the xlsx file?")

        # validate header_order
        for colorder in self.header_order:
            if colorder not in self.header_order:
                raise Exception("field '" + colorder + "' doesn't exist in header_order!")

        # validate header_order
        for row in self.data:
            for colorder in self.header_order:
               if colorder not in row:
                    raise Exception("field '" + colorder + "' doesn't exist in dict provided in constructor!")


        wb = Workbook()

        ws1 = wb.active
        ws1.title = sheet


        # make excel header
        cnt = 1
        rownumber = 1
        for colorder in self.header_order:
            _ = ws1.cell(column=cnt, row=rownumber, value=self.header_alias[colorder])
            cnt +=1

        rownumber +=1
        for row in self.data:
            cnt = 1
            for colorder in self.header_order:
                _ = ws1.cell(column=cnt, row=rownumber, value=row[colorder])
                cnt +=1
            rownumber += 1

        wb.save(filename = filename)
